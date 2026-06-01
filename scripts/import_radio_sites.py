import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import asyncio
import openpyxl
from sqlalchemy import text
from data.db import init_db, close_db, get_session_sync
from config.settings import get_settings

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "uploads", "Prédiction Budget N+1.xlsx")

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS site_tech_configs (
    site_code VARCHAR(50) PRIMARY KEY,
    has_2g BOOLEAN DEFAULT FALSE,
    has_3g BOOLEAN DEFAULT FALSE,
    has_4g_fdd BOOLEAN DEFAULT FALSE,
    has_4g_tdd BOOLEAN DEFAULT FALSE,
    has_5g BOOLEAN DEFAULT FALSE,
    radio_config VARCHAR(100),
    mes_date DATE,
    imported_at TIMESTAMP DEFAULT NOW()
);
"""

UPSERT_SQL = """
INSERT INTO site_tech_configs (site_code, has_2g, has_3g, has_4g_fdd, has_4g_tdd, has_5g, radio_config, mes_date)
VALUES (:site_code, :has_2g, :has_3g, :has_4g_fdd, :has_4g_tdd, :has_5g, :radio_config, :mes_date)
ON CONFLICT (site_code)
DO UPDATE SET
    has_2g = EXCLUDED.has_2g,
    has_3g = EXCLUDED.has_3g,
    has_4g_fdd = EXCLUDED.has_4g_fdd,
    has_4g_tdd = EXCLUDED.has_4g_tdd,
    has_5g = EXCLUDED.has_5g,
    radio_config = EXCLUDED.radio_config,
    mes_date = EXCLUDED.mes_date,
    imported_at = NOW();
"""


def _ok(val) -> bool:
    return val is not None and str(val).strip().upper() == "OK"


def read_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        code = row[0]
        if not code:
            continue
        raw_date = row[1]
        if raw_date is None:
            mes_date = None
        elif isinstance(raw_date, str):
            d = str(raw_date).strip().upper()
            mes_date = None if not d or d in ("NA", "N/A", "NONE", "") else d
        elif hasattr(raw_date, "strftime"):
            mes_date = raw_date.strftime("%Y-%m-%d")
        else:
            mes_date = str(raw_date)
        rows.append({
            "site_code": str(code).strip(),
            "has_2g": _ok(row[2]),
            "has_3g": _ok(row[3]),
            "has_4g_fdd": _ok(row[4]),
            "has_4g_tdd": _ok(row[5]),
            "has_5g": _ok(row[6]),
            "radio_config": str(row[7]).strip() if row[7] else None,
            "mes_date": mes_date,
        })
    return rows


async def main():
    settings = get_settings()
    await init_db(settings)
    session = get_session_sync()

    try:
        await session.execute(text(CREATE_TABLE_SQL))
        await session.commit()
        print("Table site_tech_configs ready.")

        rows = read_excel(EXCEL_PATH)
        print(f"Read {len(rows)} rows from Excel.")

        for i, row in enumerate(rows):
            await session.execute(text(UPSERT_SQL), row)
            if (i + 1) % 500 == 0:
                await session.commit()
                print(f"  -> {i + 1}/{len(rows)} imported")

        await session.commit()
        print(f"All {len(rows)} rows imported successfully.")

        r = await session.execute(text("SELECT COUNT(*) FROM site_tech_configs"))
        print(f"Total in table: {r.scalar()}")

    finally:
        await session.close()
        await close_db()


if __name__ == "__main__":
    asyncio.run(main())
